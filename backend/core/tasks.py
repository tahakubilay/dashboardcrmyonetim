from celery import shared_task
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import pandas as pd
import os

from .models import (
    Company, Brand, Branch, Person, Report,
    Contract, PromissoryNote, FinancialRecord
)
from .utils import (
    generate_unique_filename,
    export_financial_records_to_excel,
    export_financial_records_to_pdf,
    generate_contract_from_template
)


@shared_task(bind=True, max_retries=3)
def generate_report_task(self, user_id, scope, report_type, entity_id, **kwargs):
    """
    Ağır rapor oluşturma task'ı
    """
    try:
        user = User.objects.get(id=user_id)
        
        # Scope'a göre entity'yi al
        entity = None
        if scope == 'company':
            entity = Company.objects.get(id=entity_id)
        elif scope == 'brand':
            entity = Brand.objects.get(id=entity_id)
        elif scope == 'branch':
            entity = Branch.objects.get(id=entity_id)
        elif scope == 'person':
            entity = Person.objects.get(id=entity_id)
        
        # Rapor verilerini topla
        report_data = collect_report_data(scope, entity, report_type)
        
        # Excel dosyası oluştur
        excel_file = create_report_excel(report_data, scope, report_type)
        
        # Rapor kaydı oluştur
        report = Report.objects.create(
            title=f"{entity} - {report_type} Raporu",
            content=f"{report_type} raporu otomatik oluşturuldu",
            file=excel_file,
            report_type=report_type,
            scope=scope,
            report_date=timezone.now().date(),
            created_by=user
        )
        
        # Scope'a göre ilişkileri ayarla
        if scope == 'company':
            report.company = entity
        elif scope == 'brand':
            report.brand = entity
        elif scope == 'branch':
            report.branch = entity
        elif scope == 'person':
            report.person = entity
        
        report.save()
        
        return {
            'success': True,
            'report_id': str(report.id),
            'file_url': report.file.url if report.file else None,
            'message': 'Rapor başarıyla oluşturuldu'
        }
        
    except Exception as exc:
        # Retry on failure
        raise self.retry(exc=exc, countdown=60)


def collect_report_data(scope, entity, report_type):
    """Rapor verilerini topla"""
    data = {
        'entity': str(entity),
        'scope': scope,
        'report_type': report_type,
        'generated_at': timezone.now(),
    }
    
    # Financial records
    if scope == 'company':
        financial_records = FinancialRecord.objects.filter(related_company=entity)
    elif scope == 'brand':
        financial_records = FinancialRecord.objects.filter(related_brand=entity)
    elif scope == 'branch':
        financial_records = FinancialRecord.objects.filter(related_branch=entity)
    elif scope == 'person':
        financial_records = FinancialRecord.objects.filter(related_person=entity)
    else:
        financial_records = FinancialRecord.objects.none()
    
    # Tarih filtresi
    if report_type == 'daily':
        financial_records = financial_records.filter(date=timezone.now().date())
    elif report_type == 'weekly':
        week_ago = timezone.now().date() - timedelta(days=7)
        financial_records = financial_records.filter(date__gte=week_ago)
    elif report_type == 'monthly':
        month_ago = timezone.now().date() - timedelta(days=30)
        financial_records = financial_records.filter(date__gte=month_ago)
    elif report_type == 'yearly':
        year_ago = timezone.now().date() - timedelta(days=365)
        financial_records = financial_records.filter(date__gte=year_ago)
    
    # Özet istatistikler
    from django.db.models import Sum, Count
    
    data['statistics'] = {
        'total_records': financial_records.count(),
        'total_income': financial_records.filter(type='income').aggregate(Sum('amount'))['amount__sum'] or 0,
        'total_expense': financial_records.filter(type='expense').aggregate(Sum('amount'))['amount__sum'] or 0,
        'total_turnover': financial_records.filter(type='turnover').aggregate(Sum('amount'))['amount__sum'] or 0,
    }
    
    data['statistics']['net_profit'] = data['statistics']['total_income'] - data['statistics']['total_expense']
    
    # Detaylı kayıtlar
    data['records'] = list(financial_records.values(
        'title', 'type', 'amount', 'currency', 'date', 'description'
    ))
    
    return data


def create_report_excel(report_data, scope, report_type):
    """Rapor Excel dosyası oluştur"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, PieChart, Reference
    
    wb = Workbook()
    
    # Özet sayfası
    ws_summary = wb.active
    ws_summary.title = "Özet"
    
    # Başlık
    ws_summary['A1'] = f"{report_data['entity']} - {report_type.upper()} Raporu"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = "Rapor Tarihi:"
    ws_summary['B3'] = report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
    
    # İstatistikler
    row = 5
    ws_summary[f'A{row}'] = "İstatistikler"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    stats = report_data['statistics']
    stats_data = [
        ['Toplam Kayıt', stats['total_records']],
        ['Toplam Gelir', f"{stats['total_income']:,.2f} TL"],
        ['Toplam Gider', f"{stats['total_expense']:,.2f} TL"],
        ['Toplam Ciro', f"{stats['total_turnover']:,.2f} TL"],
        ['Net Kar', f"{stats['net_profit']:,.2f} TL"],
    ]
    
    for stat_row in stats_data:
        ws_summary[f'A{row}'] = stat_row[0]
        ws_summary[f'B{row}'] = stat_row[1]
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Detay sayfası
    if report_data['records']:
        ws_detail = wb.create_sheet("Detaylar")
        
        # Header
        headers = ['Başlık', 'Tür', 'Tutar', 'Para Birimi', 'Tarih', 'Açıklama']
        ws_detail.append(headers)
        
        # Style header
        for cell in ws_detail[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for record in report_data['records']:
            ws_detail.append([
                record['title'],
                record['type'],
                float(record['amount']),
                record['currency'],
                record['date'].strftime('%Y-%m-%d') if record['date'] else '',
                record['description'] or '',
            ])
        
        # Auto-width
        for column in ws_detail.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_detail.column_dimensions[column_letter].width = adjusted_width
    
    # Save
    filename = generate_unique_filename(f'rapor_{scope}_{report_type}', 'xlsx')
    filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    return f"reports/{filename}"


@shared_task(bind=True)
def generate_contract_pdf_task(self, contract_id):
    """Sözleşmeden PDF oluştur"""
    try:
        contract = Contract.objects.get(id=contract_id)
        
        # Sözleşme verileri
        context_data = {
            'contract_number': contract.contract_number,
            'title': contract.title,
            'start_date': contract.start_date.strftime('%Y-%m-%d'),
            'end_date': contract.end_date.strftime('%Y-%m-%d') if contract.end_date else 'Belirtilmemiş',
            'status': contract.get_status_display(),
        }
        
        # İlişkili entity bilgileri
        if contract.related_person:
            context_data['person_name'] = contract.related_person.full_name
            context_data['person_address'] = contract.related_person.address or ''
        
        if contract.related_company:
            context_data['company_name'] = contract.related_company.title
            context_data['company_tax_number'] = contract.related_company.tax_number
        
        # Template'ten oluştur
        if contract.template_name:
            doc_buffer = generate_contract_from_template(contract.template_name, context_data)
            
            # Save as PDF (docx to pdf conversion - requires additional package)
            # For now, save as docx
            filename = generate_unique_filename(f'sozlesme_{contract.contract_number}', 'docx')
            filepath = os.path.join(settings.MEDIA_ROOT, 'contracts', filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            with open(filepath, 'wb') as f:
                f.write(doc_buffer.read())
            
            # Update contract file
            contract.file = f"contracts/{filename}"
            contract.save()
        
        return {
            'success': True,
            'contract_id': str(contract.id),
            'file_url': contract.file.url if contract.file else None
        }
        
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


@shared_task
def send_expiring_contracts_notification():
    """
    Vadesi yaklaşan sözleşmeler için bildirim gönder
    Her gün çalışır (Celery Beat)
    """
    threshold_date = timezone.now().date() + timedelta(days=30)
    
    expiring_contracts = Contract.objects.filter(
        status='active',
        end_date__lte=threshold_date,
        end_date__gte=timezone.now().date()
    )
    
    if expiring_contracts.exists():
        # Admin kullanıcılara e-posta gönder
        admin_emails = User.objects.filter(is_staff=True).values_list('email', flat=True)
        
        message = f"Yakında {expiring_contracts.count()} adet sözleşmenin süresi dolacak.\n\n"
        for contract in expiring_contracts[:10]:
            message += f"- {contract.title} ({contract.contract_number}): {contract.end_date}\n"
        
        send_mail(
            subject='Vadesi Yaklaşan Sözleşmeler',
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=list(admin_emails),
            fail_silently=True,
        )
    
    return f"{expiring_contracts.count()} sözleşme için bildirim gönderildi"


@shared_task
def send_overdue_notes_notification():
    """
    Vadesi geçmiş senetler için bildirim gönder
    Her gün çalışır (Celery Beat)
    """
    overdue_notes = PromissoryNote.objects.filter(
        due_date__lt=timezone.now().date(),
        payment_status='pending'
    )
    
    if overdue_notes.exists():
        admin_emails = User.objects.filter(is_staff=True).values_list('email', flat=True)
        
        total_amount = sum(note.amount for note in overdue_notes)
        message = f"Vadesi geçmiş {overdue_notes.count()} adet senet bulunmaktadır.\n"
        message += f"Toplam Tutar: {total_amount:,.2f} TL\n\n"
        
        for note in overdue_notes[:10]:
            message += f"- {note.title} ({note.note_number}): {note.amount} TL - Vade: {note.due_date}\n"
        
        send_mail(
            subject='Vadesi Geçmiş Senetler',
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=list(admin_emails),
            fail_silently=True,
        )
    
    return f"{overdue_notes.count()} senet için bildirim gönderildi"


@shared_task
def generate_scheduled_reports():
    """
    Zamanlanmış raporları oluştur
    Her gün gece yarısı çalışır (Celery Beat)
    """
    from django.db.models import Q
    
    # Bugün oluşturulması gereken raporları kontrol et
    # (Bu özellik için ayrı bir ScheduledReport modeli gerekebilir)
    
    # Örnek: Her şirket için günlük rapor
    companies = Company.objects.filter(is_active=True)
    
    for company in companies:
        try:
            generate_report_task.delay(
                user_id=1,  # System user
                scope='company',
                report_type='daily',
                entity_id=str(company.id)
            )
        except Exception as e:
            print(f"Rapor oluşturma hatası: {e}")
    
    return f"{companies.count()} şirket için günlük rapor oluşturma başlatıldı"


@shared_task
def cleanup_old_files():
    """
    Eski dosyaları temizle
    Haftada bir çalışır
    """
    from django.core.files.storage import default_storage
    
    # 90 günden eski raporları sil
    threshold_date = timezone.now() - timedelta(days=90)
    old_reports = Report.objects.filter(created_at__lt=threshold_date)
    
    deleted_count = 0
    for report in old_reports:
        if report.file:
            try:
                default_storage.delete(report.file.name)
                deleted_count += 1
            except:
                pass
    
    old_reports.delete()
    
    return f"{deleted_count} eski dosya temizlendi"


@shared_task(bind=True, max_retries=3)
def import_excel_data_task(self, file_path, import_type, user_id):
    """
    Excel dosyasından veri import et
    import_type: 'financial_records', 'people', vb.
    """
    try:
        user = User.objects.get(id=user_id)
        
        if import_type == 'financial_records':
            from .utils import import_financial_records_from_excel
            result = import_financial_records_from_excel(file_path, user)
            return result
        
        return {'success': False, 'error': 'Geçersiz import tipi'}
        
    except Exception as exc:
        raise self.retry(exc=exc, countdown=60)


@shared_task
def update_contract_statuses():
    """
    Sözleşme durumlarını güncelle (expired olarak işaretle)
    Her gün çalışır
    """
    today = timezone.now().date()
    
    # Süreleri dolan sözleşmeleri expired yap
    expired_count = Contract.objects.filter(
        status='active',
        end_date__lt=today
    ).update(status='expired')
    
    return f"{expired_count} sözleşme expired olarak işaretlendi"


@shared_task
def update_promissory_note_statuses():
    """
    Senet durumlarını güncelle (overdue olarak işaretle)
    Her gün çalışır
    """
    today = timezone.now().date()
    
    # Vadesi geçen senetleri overdue yap
    overdue_count = PromissoryNote.objects.filter(
        payment_status='pending',
        due_date__lt=today
    ).update(payment_status='overdue')
    
    return f"{overdue_count} senet overdue olarak işaretlendi"


@shared_task
def generate_monthly_financial_summary():
    """
    Aylık mali özet raporu oluştur
    Her ayın 1'inde çalışır
    """
    from django.db.models import Sum
    from dateutil.relativedelta import relativedelta
    
    # Geçen ay
    today = timezone.now().date()
    last_month_start = (today.replace(day=1) - relativedelta(months=1))
    last_month_end = today.replace(day=1) - timedelta(days=1)
    
    # Her şirket için özet
    companies = Company.objects.filter(is_active=True)
    
    for company in companies:
        records = FinancialRecord.objects.filter(
            related_company=company,
            date__gte=last_month_start,
            date__lte=last_month_end
        )
        
        summary = {
            'company': company.title,
            'period': f"{last_month_start} - {last_month_end}",
            'total_income': records.filter(type='income').aggregate(Sum('amount'))['amount__sum'] or 0,
            'total_expense': records.filter(type='expense').aggregate(Sum('amount'))['amount__sum'] or 0,
            'total_turnover': records.filter(type='turnover').aggregate(Sum('amount'))['amount__sum'] or 0,
        }
        
        summary['net_profit'] = summary['total_income'] - summary['total_expense']
        
        # E-posta gönder
        admin_emails = User.objects.filter(is_staff=True).values_list('email', flat=True)
        
        message = f"Aylık Mali Özet - {company.title}\n"
        message += f"Dönem: {summary['period']}\n\n"
        message += f"Toplam Gelir: {summary['total_income']:,.2f} TL\n"
        message += f"Toplam Gider: {summary['total_expense']:,.2f} TL\n"
        message += f"Toplam Ciro: {summary['total_turnover']:,.2f} TL\n"
        message += f"Net Kar: {summary['net_profit']:,.2f} TL\n"
        
        send_mail(
            subject=f'Aylık Mali Özet - {company.title}',
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=list(admin_emails),
            fail_silently=True,
        )
    
    return f"{companies.count()} şirket için aylık özet gönderildi"


@shared_task
def backup_database():
    """
    Veritabanı yedeği al (PostgreSQL için)
    Haftada bir çalışır
    """
    import subprocess
    from django.conf import settings
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = f"backup_{timestamp}.sql"
    backup_path = os.path.join(settings.MEDIA_ROOT, 'backups', backup_file)
    os.makedirs(os.path.dirname(backup_path), exist_ok=True)
    
    try:
        # PostgreSQL backup
        db_config = settings.DATABASES['default']
        cmd = [
            'pg_dump',
            '-h', db_config['HOST'],
            '-U', db_config['USER'],
            '-d', db_config['NAME'],
            '-f', backup_path
        ]
        
        env = os.environ.copy()
        env['PGPASSWORD'] = db_config['PASSWORD']
        
        subprocess.run(cmd, env=env, check=True)
        
        # S3'e yükle (opsiyonel)
        if settings.USE_S3:
            # Upload to S3
            pass
        
        return f"Yedek başarıyla alındı: {backup_file}"
        
    except Exception as e:
        return f"Yedek hatası: {str(e)}"